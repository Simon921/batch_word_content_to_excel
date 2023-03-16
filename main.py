import os
import time
import tkinter as tk
from tkinter import filedialog
import docx
import openpyxl


def get_word_files(directory):
    """获取目录中扩展名为 .docx 的 Word 文件"""
    files = []
    for filename in os.listdir(directory):
        if filename.endswith('.docx'):
            files.append(os.path.join(directory, filename))
    return files


def extract_content_from_word(file):
    """从 Word 文件中提取文件名和正文内容"""
    doc = docx.Document(file)
    filename = os.path.basename(file)
    content = ''
    for para in doc.paragraphs:
        content += para.text + '\n'
    return filename, content


def save_to_excel(data, output_file):
    """将数据保存到 Excel 文件中"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = time.strftime("%Y%m%d", time.localtime()) + "-Word内容提取"
    ws['A1'] = "文件名"
    ws['B1'] = "正文"
    for row in data:
        ws.append(row)
    wb.save(output_file)


def select_directory():
    """选择要提取文件的目录"""
    directory = filedialog.askdirectory()
    if directory:
        extract_word_content(directory)


def extract_word_content(directory):
    """批量提取 Word 文件的文件名和正文内容"""
    files = get_word_files(directory)
    data = []
    for file in files:
        filename, content = extract_content_from_word(file)
        data.append((filename, content))
    output_file = os.path.join(directory, time.strftime("%Y%m%d", time.localtime()) + " Word内容提取.xlsx")
    save_to_excel(data, output_file)
    message_label.config(text=f"提取完成，已保存到 {output_file}")


# 创建 GUI 界面
window = tk.Tk()
window.title("批量提取Word文件名+正文")
window.geometry("600x100")

# 添加按钮和标签
select_button = tk.Button(text="选择文件夹", command=select_directory)
select_button.pack(pady=20)
message_label = tk.Label(text="")
message_label.pack()

window.mainloop()
